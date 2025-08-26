from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import os
import json
from datetime import datetime, timedelta
import requests
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Data directory path
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')

def ensure_excel_file(filename, headers):
    """Ensure Excel file exists with proper headers"""
    filepath = os.path.join(DATA_DIR, filename)
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        wb.save(filepath)
        logger.info(f"Created {filename} with headers: {headers}")
    return filepath

def read_excel_data(filename):
    """Read data from Excel file"""
    try:
        filepath = os.path.join(DATA_DIR, filename)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        # Get data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
        
        return data
    except Exception as e:
        logger.error(f"Error reading {filename}: {str(e)}")
        return []

def write_excel_data(filename, data, headers):
    """Write data to Excel file"""
    try:
        filepath = os.path.join(DATA_DIR, filename)
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        logger.info(f"Successfully wrote data to {filename}")
        return True
    except Exception as e:
        logger.error(f"Error writing to {filename}: {str(e)}")
        return False

def get_next_id(data):
    """Get next available ID"""
    if not data:
        return 1
    max_id = max([int(item.get('id', 0)) for item in data if item.get('id')])
    return max_id + 1

# Initialize Excel files
@app.before_first_request
def initialize_files():
    """Initialize all Excel files with proper headers"""
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # Patients file
    ensure_excel_file('patients.xlsx', [
        'id', 'name', 'age', 'gender', 'phone', 'email', 'address', 
        'medical_history', 'allergies', 'emergency_contact', 'created_date'
    ])
    
    # Appointments file
    ensure_excel_file('appointments.xlsx', [
        'id', 'patient_id', 'patient_name', 'appointment_date', 'appointment_time',
        'status', 'reason', 'notes', 'created_date'
    ])
    
    # Billing file
    ensure_excel_file('billing.xlsx', [
        'id', 'patient_id', 'patient_name', 'service_description', 'amount',
        'status', 'payment_method', 'bill_date', 'due_date', 'created_date'
    ])
    
    # Reports file
    ensure_excel_file('reports.xlsx', [
        'id', 'patient_id', 'patient_name', 'report_type', 'diagnosis',
        'treatment', 'medications', 'follow_up_date', 'created_date'
    ])
    
    # Prescriptions file
    ensure_excel_file('prescriptions.xlsx', [
        'id', 'patient_id', 'patient_name', 'medication_name', 'dosage',
        'frequency', 'duration', 'instructions', 'prescribed_date'
    ])

# PATIENTS ENDPOINTS
@app.route('/patients', methods=['GET'])
def get_patients():
    """Get all patients"""
    try:
        data = read_excel_data('patients.xlsx')
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/patients', methods=['POST'])
def add_patient():
    """Add new patient"""
    try:
        patient_data = request.json
        existing_data = read_excel_data('patients.xlsx')
        
        # Add ID and created date
        patient_data['id'] = get_next_id(existing_data)
        patient_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        existing_data.append(patient_data)
        
        headers = ['id', 'name', 'age', 'gender', 'phone', 'email', 'address', 
                  'medical_history', 'allergies', 'emergency_contact', 'created_date']
        
        if write_excel_data('patients.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Patient added successfully', 'data': patient_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to save patient data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/patients/<int:patient_id>', methods=['PUT'])
def update_patient(patient_id):
    """Update patient"""
    try:
        update_data = request.json
        existing_data = read_excel_data('patients.xlsx')
        
        # Find and update patient
        updated = False
        for i, patient in enumerate(existing_data):
            if int(patient.get('id', 0)) == patient_id:
                existing_data[i].update(update_data)
                updated = True
                break
        
        if not updated:
            return jsonify({'success': False, 'error': 'Patient not found'}), 404
        
        headers = ['id', 'name', 'age', 'gender', 'phone', 'email', 'address', 
                  'medical_history', 'allergies', 'emergency_contact', 'created_date']
        
        if write_excel_data('patients.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Patient updated successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update patient data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/patients/<int:patient_id>', methods=['DELETE'])
def delete_patient(patient_id):
    """Delete patient"""
    try:
        existing_data = read_excel_data('patients.xlsx')
        
        # Filter out the patient to delete
        filtered_data = [p for p in existing_data if int(p.get('id', 0)) != patient_id]
        
        if len(filtered_data) == len(existing_data):
            return jsonify({'success': False, 'error': 'Patient not found'}), 404
        
        headers = ['id', 'name', 'age', 'gender', 'phone', 'email', 'address', 
                  'medical_history', 'allergies', 'emergency_contact', 'created_date']
        
        if write_excel_data('patients.xlsx', filtered_data, headers):
            return jsonify({'success': True, 'message': 'Patient deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to delete patient'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# APPOINTMENTS ENDPOINTS
@app.route('/appointments', methods=['GET'])
def get_appointments():
    """Get all appointments"""
    try:
        data = read_excel_data('appointments.xlsx')
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/appointments', methods=['POST'])
def add_appointment():
    """Add new appointment"""
    try:
        appointment_data = request.json
        existing_data = read_excel_data('appointments.xlsx')
        
        # Add ID and created date
        appointment_data['id'] = get_next_id(existing_data)
        appointment_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        appointment_data['status'] = appointment_data.get('status', 'Pending')
        
        existing_data.append(appointment_data)
        
        headers = ['id', 'patient_id', 'patient_name', 'appointment_date', 'appointment_time',
                  'status', 'reason', 'notes', 'created_date']
        
        if write_excel_data('appointments.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Appointment added successfully', 'data': appointment_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to save appointment data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/appointments/<int:appointment_id>', methods=['PUT'])
def update_appointment(appointment_id):
    """Update appointment (for confirm/reschedule)"""
    try:
        update_data = request.json
        existing_data = read_excel_data('appointments.xlsx')
        
        # Find and update appointment
        updated = False
        for i, appointment in enumerate(existing_data):
            if int(appointment.get('id', 0)) == appointment_id:
                existing_data[i].update(update_data)
                updated = True
                break
        
        if not updated:
            return jsonify({'success': False, 'error': 'Appointment not found'}), 404
        
        headers = ['id', 'patient_id', 'patient_name', 'appointment_date', 'appointment_time',
                  'status', 'reason', 'notes', 'created_date']
        
        if write_excel_data('appointments.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Appointment updated successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update appointment'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/appointments/<int:appointment_id>', methods=['DELETE'])
def delete_appointment(appointment_id):
    """Delete appointment"""
    try:
        existing_data = read_excel_data('appointments.xlsx')
        
        # Filter out the appointment to delete
        filtered_data = [a for a in existing_data if int(a.get('id', 0)) != appointment_id]
        
        if len(filtered_data) == len(existing_data):
            return jsonify({'success': False, 'error': 'Appointment not found'}), 404
        
        headers = ['id', 'patient_id', 'patient_name', 'appointment_date', 'appointment_time',
                  'status', 'reason', 'notes', 'created_date']
        
        if write_excel_data('appointments.xlsx', filtered_data, headers):
            return jsonify({'success': True, 'message': 'Appointment deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to delete appointment'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# BILLING ENDPOINTS
@app.route('/billing', methods=['GET'])
def get_billing():
    """Get all billing records"""
    try:
        data = read_excel_data('billing.xlsx')
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/billing', methods=['POST'])
def add_billing():
    """Add new billing record"""
    try:
        billing_data = request.json
        existing_data = read_excel_data('billing.xlsx')
        
        # Add ID and created date
        billing_data['id'] = get_next_id(existing_data)
        billing_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        billing_data['status'] = billing_data.get('status', 'Pending')
        
        existing_data.append(billing_data)
        
        headers = ['id', 'patient_id', 'patient_name', 'service_description', 'amount',
                  'status', 'payment_method', 'bill_date', 'due_date', 'created_date']
        
        if write_excel_data('billing.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Billing record added successfully', 'data': billing_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to save billing data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# REPORTS ENDPOINTS
@app.route('/reports', methods=['GET'])
def get_reports():
    """Get all reports"""
    try:
        data = read_excel_data('reports.xlsx')
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/reports', methods=['POST'])
def add_report():
    """Add new report"""
    try:
        report_data = request.json
        existing_data = read_excel_data('reports.xlsx')
        
        # Add ID and created date
        report_data['id'] = get_next_id(existing_data)
        report_data['created_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        existing_data.append(report_data)
        
        headers = ['id', 'patient_id', 'patient_name', 'report_type', 'diagnosis',
                  'treatment', 'medications', 'follow_up_date', 'created_date']
        
        if write_excel_data('reports.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Report added successfully', 'data': report_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to save report data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# PRESCRIPTIONS ENDPOINTS
@app.route('/prescriptions', methods=['GET'])
def get_prescriptions():
    """Get all prescriptions"""
    try:
        data = read_excel_data('prescriptions.xlsx')
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/prescriptions', methods=['POST'])
def add_prescription():
    """Add new prescription"""
    try:
        prescription_data = request.json
        existing_data = read_excel_data('prescriptions.xlsx')
        
        # Add ID and prescribed date
        prescription_data['id'] = get_next_id(existing_data)
        prescription_data['prescribed_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        existing_data.append(prescription_data)
        
        headers = ['id', 'patient_id', 'patient_name', 'medication_name', 'dosage',
                  'frequency', 'duration', 'instructions', 'prescribed_date']
        
        if write_excel_data('prescriptions.xlsx', existing_data, headers):
            return jsonify({'success': True, 'message': 'Prescription added successfully', 'data': prescription_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to save prescription data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ANALYTICS ENDPOINT
@app.route('/analytics', methods=['GET'])
def get_analytics():
    """Get analytics data"""
    try:
        appointments = read_excel_data('appointments.xlsx')
        patients = read_excel_data('patients.xlsx')
        billing = read_excel_data('billing.xlsx')
        
        # Calculate metrics
        total_appointments = len(appointments)
        pending_billing = len([b for b in billing if b.get('status') == 'Pending'])
        active_patients = len(patients)
        
        # Today's appointments
        today = datetime.now().strftime('%Y-%m-%d')
        todays_appointments = [a for a in appointments if a.get('appointment_date') == today]
        
        analytics_data = {
            'total_appointments': total_appointments,
            'pending_billing': pending_billing,
            'active_patients': active_patients,
            'avg_wait_time': '15 mins',  # Placeholder
            'todays_appointments': todays_appointments,
            'revenue_trend': 'Increasing',  # Placeholder
            'follow_up_alerts': len([a for a in appointments if a.get('status') == 'Pending'])
        }
        
        return jsonify({'success': True, 'data': analytics_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# NOTIFICATIONS ENDPOINT (Dummy WhatsApp API)
@app.route('/notify', methods=['POST'])
def send_notification():
    """Send WhatsApp notification (dummy implementation)"""
    try:
        notification_data = request.json
        message = notification_data.get('message', '')
        patient_phone = notification_data.get('phone', '')
        
        # Dummy WhatsApp API call
        logger.info(f"Sending WhatsApp notification to {patient_phone}: {message}")
        
        # Simulate API response
        response = {
            'success': True,
            'message': 'Notification sent successfully',
            'whatsapp_response': {
                'message_id': f"msg_{datetime.now().timestamp()}",
                'status': 'sent',
                'phone': patient_phone
            }
        }
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
